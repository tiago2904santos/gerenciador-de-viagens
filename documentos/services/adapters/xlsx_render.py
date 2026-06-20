"""Preenchimento de planilhas-modelo (.xlsx) via openpyxl.

O modelo do diário de bordo (``documentos/resources/diario_bordo.xlsx``) tem
placeholders ``{{...}}`` no cabeçalho/rodapé e **uma** linha de trecho (a 17ª)
que serve de gabarito. Como ``insert_rows`` do openpyxl não reposiciona merges
nem copia estilos, fazemos a expansão de linhas manualmente: clonamos o estilo
da linha-gabarito para cada trecho e reescrevemos o rodapé deslocado.
"""

from __future__ import annotations

import re
from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Mapping
from typing import Sequence

_PLACEHOLDER_RE = re.compile(r"{{\s*([\w.]+)\s*}}")

# Layout do modelo diario_bordo.xlsx (1-indexed).
TRECHO_ROW = 17           # linha-gabarito com os placeholders de trecho
FOOTER_FIRST_ROW = 18     # primeira linha do rodapé (NOME / RG / ASSINATURA)
FOOTER_LAST_ROW = 20      # última linha do rodapé
LAST_COL = 12             # coluna L

# Colunas (1-indexed) dos campos de cada trecho. Origem/Destino/Abastecimento
# são células mescladas (G:H, I:J, K:L); preenchemos a célula-âncora.
TRECHO_COLUNAS = {
    "data_saida": 1,    # A
    "hora_saida": 2,    # B
    "km_inicial": 3,    # C
    "data_chegada": 4,  # D
    "hora_chegada": 5,  # E
    "km_final": 6,      # F
    "origem": 7,        # G (G:H)
    "destino": 9,       # I (I:J)
    "abastecimento": 11,  # K (K:L)
}


def _resolver_placeholders(texto: str, contexto: Mapping[str, object]) -> str:
    def _sub(match: "re.Match[str]") -> str:
        chave = match.group(1)
        if chave in contexto:
            return str(contexto[chave] if contexto[chave] is not None else "")
        return match.group(0)

    return _PLACEHOLDER_RE.sub(_sub, texto)


def _copiar_estilo(origem, destino) -> None:
    destino.font = copy(origem.font)
    destino.border = copy(origem.border)
    destino.fill = copy(origem.fill)
    destino.alignment = copy(origem.alignment)
    destino.number_format = origem.number_format
    destino.protection = copy(origem.protection)


def _merges_na_linha(ws, linha: int) -> list[tuple[int, int]]:
    """Retorna (col_inicial, col_final) das mesclagens contidas na ``linha``."""
    spans = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row == linha and rng.max_row == linha:
            spans.append((rng.min_col, rng.max_col))
    return spans


def fill_diario_bordo_xlsx(
    *,
    template_path: Path,
    header: Mapping[str, object],
    trechos: Sequence[Mapping[str, object]],
) -> bytes:
    """Preenche o modelo do diário de bordo e devolve os bytes do .xlsx.

    ``header`` mapeia placeholders simples ({{divisao}}, {{placa}}, {{motorista}}…).
    ``trechos`` é a lista de linhas; cada item usa as chaves de ``TRECHO_COLUNAS``.
    """
    try:
        import openpyxl
    except ModuleNotFoundError as exc:  # pragma: no cover - dependência opcional
        raise RuntimeError(
            "Pacote 'openpyxl' não instalado no ambiente ativo. "
            "Execute: python -m pip install openpyxl",
        ) from exc

    wb = openpyxl.load_workbook(str(template_path))
    ws = wb.active

    n = max(len(trechos), 1)

    # 1) Captura o gabarito do trecho (estilos + mesclagens internas da linha).
    trecho_spans = _merges_na_linha(ws, TRECHO_ROW)
    trecho_estilos = {
        col: copy(ws.cell(row=TRECHO_ROW, column=col)._style) for col in range(1, LAST_COL + 1)
    }
    trecho_altura = ws.row_dimensions[TRECHO_ROW].height

    # 2) Captura o rodapé (valores + estilos + mesclagens) e suas alturas.
    footer_spans = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= FOOTER_FIRST_ROW and rng.max_row <= FOOTER_LAST_ROW:
            footer_spans.append(
                (rng.min_row - FOOTER_FIRST_ROW, rng.max_row - FOOTER_FIRST_ROW, rng.min_col, rng.max_col),
            )
    footer_cells = []  # (offset_linha, col, valor, style)
    for r in range(FOOTER_FIRST_ROW, FOOTER_LAST_ROW + 1):
        for col in range(1, LAST_COL + 1):
            cell = ws.cell(row=r, column=col)
            footer_cells.append((r - FOOTER_FIRST_ROW, col, cell.value, copy(cell._style)))
    footer_alturas = {
        r - FOOTER_FIRST_ROW: ws.row_dimensions[r].height for r in range(FOOTER_FIRST_ROW, FOOTER_LAST_ROW + 1)
    }

    # 3) Remove mesclagens do gabarito e do rodapé antes de reescrever.
    for rng in list(ws.merged_cells.ranges):
        if (rng.min_row == TRECHO_ROW and rng.max_row == TRECHO_ROW) or (
            rng.min_row >= FOOTER_FIRST_ROW and rng.max_row <= FOOTER_LAST_ROW
        ):
            ws.unmerge_cells(str(rng))

    # 4) Limpa a região antiga (gabarito + rodapé) — será toda reescrita.
    for r in range(TRECHO_ROW, FOOTER_LAST_ROW + 1):
        for col in range(1, LAST_COL + 1):
            ws.cell(row=r, column=col).value = None

    # 5) Escreve uma linha por trecho clonando o estilo do gabarito.
    for i in range(n):
        linha = TRECHO_ROW + i
        for col in range(1, LAST_COL + 1):
            ws.cell(row=linha, column=col)._style = copy(trecho_estilos[col])
        if trecho_altura is not None:
            ws.row_dimensions[linha].height = trecho_altura
        dados = trechos[i] if i < len(trechos) else {}
        for chave, col in TRECHO_COLUNAS.items():
            valor = dados.get(chave)
            ws.cell(row=linha, column=col).value = "" if valor is None else valor
        for c0, c1 in trecho_spans:
            ws.merge_cells(start_row=linha, start_column=c0, end_row=linha, end_column=c1)

    # 6) Reescreve o rodapé deslocado para depois do último trecho.
    footer_base = TRECHO_ROW + n
    for offset, col, valor, style in footer_cells:
        cell = ws.cell(row=footer_base + offset, column=col)
        cell._style = copy(style)
        if isinstance(valor, str) and "{{" in valor:
            cell.value = _resolver_placeholders(valor, header)
        else:
            cell.value = valor
    for offset, altura in footer_alturas.items():
        if altura is not None:
            ws.row_dimensions[footer_base + offset].height = altura
    for r0, r1, c0, c1 in footer_spans:
        ws.merge_cells(
            start_row=footer_base + r0,
            start_column=c0,
            end_row=footer_base + r1,
            end_column=c1,
        )

    # 7) Resolve os placeholders do cabeçalho (tudo acima da linha-gabarito).
    for row in ws.iter_rows(min_row=1, max_row=TRECHO_ROW - 1):
        for cell in row:
            if isinstance(cell.value, str) and "{{" in cell.value:
                cell.value = _resolver_placeholders(cell.value, header)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
